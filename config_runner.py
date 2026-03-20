#!/usr/bin/env python3
"""
【設定駆動型ランナー】
config.json だけで全ての処理が変わる

• 企業リストは config.json で管理
• Strategy の重み付けも config.json で変更可能
• 出力形式も config.json で指定
• Null値やエラーも柔軟に処理

プログラムの修正は一切不要！
"""

import json
import logging
import os
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path
import sys

# ========== 【ステップ1】設定ファイルの読み込み ==========

class ConfigLoader:
    """設定ファイルを読み込んで検証"""
    
    @staticmethod
    def load(config_file: str = 'config_phase1.json') -> Dict:
        """
        設定ファイルを読み込む
        
        Returns:
            config: 設定辞書
        """
        if not os.path.exists(config_file):
            raise FileNotFoundError(f"設定ファイルが見つかりません: {config_file}")
        
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        return config
    
    @staticmethod
    def validate(config: Dict) -> bool:
        """
        設定の妥当性を検証
        
        必須フィールド：
        • phase.current
        • data_sources.jquants
        • companies.data
        • strategies
        """
        required_fields = [
            'phase',
            'data_sources',
            'companies',
            'strategies'
        ]
        
        for field in required_fields:
            if field not in config:
                raise ValueError(f"必須フィールドが見つかりません: {field}")
        
        return True

# ========== 【ステップ2】ログシステムの初期化 ==========

class LoggerSetup:
    """ログシステムの初期化"""
    
    @staticmethod
    def setup(config: Dict) -> logging.Logger:
        """
        設定ファイルに基づいてログシステムを初期化
        """
        log_config = config.get('logging', {})
        level = getattr(logging, log_config.get('level', 'INFO'))
        
        # ログディレクトリを作成
        log_dir = os.path.dirname(log_config.get('file_path', './logs/app.log'))
        os.makedirs(log_dir, exist_ok=True)
        
        # ログシステムを設定
        logging.basicConfig(
            level=level,
            format=log_config.get('format', '%(asctime)s [%(levelname)s] %(name)s: %(message)s'),
            handlers=[
                logging.FileHandler(log_config.get('file_path', './logs/app.log')),
                logging.StreamHandler(sys.stdout)
            ]
        )
        
        logger = logging.getLogger(__name__)
        logger.info(f"ログシステムを初期化しました (Level: {log_config.get('level', 'INFO')})")
        
        return logger

# ========== 【ステップ3】企業データの抽出 ==========

class CompanyExtractor:
    """
    設定ファイルから企業データを抽出
    
    • Phase に応じてフィルタリング
    • Null値を許容
    • エラーを継続
    """
    
    @staticmethod
    def extract_for_phase(config: Dict, logger: logging.Logger) -> List[Dict]:
        """
        現在の Phase に該当する企業を抽出
        """
        current_phase = config['phase']['current']
        
        # Phase 設定から企業数を取得
        phase_config = config['phase']['phases'].get(current_phase, {})
        max_companies = phase_config.get('num_companies', 15)
        
        logger.info(f"📋 {current_phase} のデータを抽出中...")
        logger.info(f"   目標企業数: {max_companies}社")
        logger.info(f"   目的: {', '.join(phase_config.get('objective', []))}")
        
        # 企業データを取得
        companies = config.get('companies', {}).get('data', [])
        
        # Phase フィルタを適用
        phase_key = current_phase.replace('_', '')  # phase_1 → phase1
        filtered = [
            c for c in companies
            if c.get(current_phase, True) is True
        ][:max_companies]
        
        logger.info(f"✅ {len(filtered)}社を抽出しました")
        
        return filtered

# ========== 【ステップ4】企業データのフォーマット変換 ==========

class DataFormatter:
    """
    企業データを異なるフォーマットに自動変換
    プログラム修正なし！
    """
    
    @staticmethod
    def to_list(companies: List[Dict]) -> List[str]:
        """企業コードのリスト化"""
        return [c['code'] for c in companies]
    
    @staticmethod
    def to_dict_by_code(companies: List[Dict]) -> Dict[str, Dict]:
        """企業コードをキーにした辞書化"""
        return {c['code']: c for c in companies}
    
    @staticmethod
    def to_csv(companies: List[Dict], output_path: str = './results/companies.csv'):
        """CSV 化"""
        import csv
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['code', 'name', 'sector'])
            writer.writeheader()
            for c in companies:
                writer.writerow({
                    'code': c['code'],
                    'name': c['name'],
                    'sector': c['sector']
                })
        
        return output_path
    
    @staticmethod
    def to_excel(companies: List[Dict], output_path: str = './results/companies.xlsx'):
        """Excel 化"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Companies'
            
            # ヘッダー
            headers = ['Code', 'Name', 'Sector', 'Dividend Yield', 'EPS Growth', 'PER', 'PBR']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            
            # データ行
            for row_idx, company in enumerate(companies, 2):
                ws.cell(row=row_idx, column=1, value=company['code'])
                ws.cell(row=row_idx, column=2, value=company['name'])
                ws.cell(row=row_idx, column=3, value=company['sector'])
                
                # Metadata から値を取得（Null許容）
                metadata = company.get('metadata', {})
                ws.cell(row=row_idx, column=4, value=metadata.get('dividend_yield'))
                ws.cell(row=row_idx, column=5, value=metadata.get('eps_growth'))
                ws.cell(row=row_idx, column=6, value=metadata.get('per'))
                ws.cell(row=row_idx, column=7, value=metadata.get('pbr'))
            
            # 列幅調整
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            
            wb.save(output_path)
            
            return output_path
        except ImportError:
            logging.warning("openpyxl がインストールされていません。Excel 化をスキップします")
            return None
    
    @staticmethod
    def to_json(companies: List[Dict], output_path: str = './results/companies.json'):
        """JSON 化"""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(companies, f, ensure_ascii=False, indent=2)
        
        return output_path

# ========== 【ステップ5】バリデーション ==========

class Validator:
    """
    Phase 1 の期待結果とのバリデーション
    """
    
    @staticmethod
    def print_expectations(config: Dict, logger: logging.Logger):
        """期待結果を表示"""
        validation = config.get('validation', {}).get('phase_1', {})
        expected = validation.get('expected_results', {})
        
        logger.info("\n" + "=" * 70)
        logger.info("【期待される結果】")
        logger.info("=" * 70)
        
        # Dividend Strategy
        div_top1 = expected.get('dividend_strategy', {}).get('top_1', {})
        logger.info(f"\nDividend Strategy:")
        logger.info(f"  1位: {div_top1.get('code')} {div_top1.get('name')} - {div_top1.get('reason')}")
        logger.info(f"  Top 3: {expected.get('dividend_strategy', {}).get('top_3_codes', [])}")
        
        # Growth Strategy
        gr_top1 = expected.get('growth_strategy', {}).get('top_1', {})
        logger.info(f"\nGrowth Strategy:")
        logger.info(f"  1位: {gr_top1.get('code')} {gr_top1.get('name')} - {gr_top1.get('reason')}")
        logger.info(f"  Top 3: {expected.get('growth_strategy', {}).get('top_3_codes', [])}")
        
        # Value Strategy
        logger.info(f"\nValue Strategy:")
        logger.info(f"  下位 3: {expected.get('value_strategy', {}).get('bottom_3_codes', [])}")
        logger.info(f"  理由: {expected.get('value_strategy', {}).get('reason', '')}")
        
        # チェックリスト
        logger.info(f"\n【チェック項目】")
        for check in validation.get('checks', []):
            logger.info(f"  □ {check}")

# ========== 【ステップ6】メイン処理 ==========

class ConfigDrivenRunner:
    """設定駆動型ランナー"""
    
    def __init__(self, config_file: str = 'config_phase1.json'):
        self.config_file = config_file
        self.config = None
        self.logger = None
        self.companies = None
    
    def run(self):
        """
        一連の処理を実行
        """
        print("\n" + "=" * 70)
        print("【設定駆動型ランナー】")
        print("=" * 70 + "\n")
        
        # ステップ 1: 設定ファイルの読み込み
        print("✓ ステップ 1: 設定ファイルを読み込み中...")
        self.config = ConfigLoader.load(self.config_file)
        ConfigLoader.validate(self.config)
        print(f"  ✅ {self.config_file} を読み込みました\n")
        
        # ステップ 2: ログシステムの初期化
        print("✓ ステップ 2: ログシステムを初期化中...")
        self.logger = LoggerSetup.setup(self.config)
        print(f"  ✅ ログシステムを初期化しました\n")
        
        # ステップ 3: 企業データの抽出
        print("✓ ステップ 3: 企業データを抽出中...")
        self.companies = CompanyExtractor.extract_for_phase(self.config, self.logger)
        print(f"  ✅ {len(self.companies)}社を抽出しました\n")
        
        # ステップ 4: データフォーマット変換
        print("✓ ステップ 4: データを自動変換中...")
        self._convert_formats()
        
        # ステップ 5: バリデーション表示
        print("\n✓ ステップ 5: バリデーション基準を表示中...")
        Validator.print_expectations(self.config, self.logger)
        
        print("\n" + "=" * 70)
        print("✅ 準備完了！")
        print("=" * 70)
        print(f"\n【次のステップ】")
        print(f"  1. advanced_architecture.py で Phase 1 を実行")
        print(f"  2. 期待される結果との比較をする")
        print(f"  3. OK なら Phase 2 へ")
        
        return self.companies
    
    def _convert_formats(self):
        """各フォーマットへの変換を実行"""
        output_config = self.config.get('output', {})
        
        # リスト化
        companies_list = DataFormatter.to_list(self.companies)
        self.logger.info(f"\n📋 企業コードリスト（{len(companies_list)}社）:")
        self.logger.info(f"   {companies_list}")
        print(f"  ✓ リスト化: {len(companies_list)} 企業")
        
        # 辞書化
        companies_dict = DataFormatter.to_dict_by_code(self.companies)
        self.logger.info(f"\n📋 辞書化（キー: 企業コード）: {len(companies_dict)} 社")
        print(f"  ✓ 辞書化: {len(companies_dict)} 企業")
        
        # CSV
        if output_config.get('csv', {}).get('enabled', False):
            csv_path = DataFormatter.to_csv(
                self.companies,
                output_config['csv'].get('output_path', './results/companies.csv')
            )
            self.logger.info(f"\n📁 CSV 出力: {csv_path}")
            print(f"  ✓ CSV: {csv_path}")
        
        # Excel
        if output_config.get('excel', {}).get('enabled', False):
            excel_path = DataFormatter.to_excel(
                self.companies,
                output_config['excel'].get('output_path', './results/companies.xlsx')
            )
            if excel_path:
                self.logger.info(f"\n📊 Excel 出力: {excel_path}")
                print(f"  ✓ Excel: {excel_path}")
        
        # JSON
        if output_config.get('json', {}).get('enabled', False):
            json_path = DataFormatter.to_json(
                self.companies,
                output_config['json'].get('output_path', './results/companies.json')
            )
            self.logger.info(f"\n📄 JSON 出力: {json_path}")
            print(f"  ✓ JSON: {json_path}")

# ========== 【実行】 ==========

if __name__ == '__main__':
    runner = ConfigDrivenRunner('config_phase1.json')
    companies = runner.run()
    
    print("\n" + "=" * 70)
    print("【抽出されたデータ】")
    print("=" * 70)
    print(f"\n企業数: {len(companies)}社\n")
    
    for i, company in enumerate(companies, 1):
        metadata = company.get('metadata', {})
        
        # Null値を許容し、表示時に "N/A" に変換
        dividend = metadata.get('dividend_yield')
        growth = metadata.get('eps_growth')
        per = metadata.get('per')
        
        dividend_str = f"{dividend:>5}" if dividend is not None else "  N/A"
        growth_str = f"{growth:>6}" if growth is not None else "   N/A"
        per_str = f"{per:>6}" if per is not None else "   N/A"
        
        print(f"{i:2d}. {company['code']} | {company['name']:15s} | 配当: {dividend_str}, "
              f"成長: {growth_str}, PER: {per_str}")
    
    print("\n" + "=" * 70)
    print("✅ config_phase1.json だけを編集することで、全ての処理が変わります")
    print("   プログラムの修正は一切不要！")
    print("=" * 70)
